import cfscrape
import json
from time import sleep
import openpyxl
from openpyxl.styles import PatternFill
import os

list_value = 0
excelFileName = 'csgorun_statistic_1.xlsx'
excelFileName_remove = 'csgorun_statistic_1.xlsx'
color_value = True
column = 2
last_game_id = 0


while True:
    scraper = cfscrape.create_scraper()
    raw_json_data = json.loads(scraper.get("https://api.csgorun.org/current-state").content)

    date = raw_json_data['date']
    data_count = raw_json_data['data']['game']['statistic']['count']
    data_totalDeposit = raw_json_data['data']['game']['statistic']['totalDeposit']
    data_totalItems = raw_json_data['data']['game']['statistic']['totalItems']

    history = raw_json_data['data']['game']['history']
    history_id = history[0]['id']
    history_crash = history[0]['crash']

    print('date: ', date)
    print('id: ', history_id)
    print('count: ', data_count)
    print('totalItems: ', data_totalItems)
    print('totalDeposit: ', data_totalDeposit)
    print('crash: ', history_crash)
    print('\n', '##############################', '\n')

    # всё в эксель, жду код котик <3
    if history_id > last_game_id:

        if list_value == 0:
            wb = openpyxl.load_workbook(filename = 'csgorun_statistic_basic.xlsx')
        else:
            wb = openpyxl.load_workbook(excelFileName)
            os.remove(excelFileName_remove)
        sheet = wb['main']

        if color_value:
            sheet['A' + str(column)] = date
            sheet['A' + str(column)].fill = PatternFill(fill_type='solid', start_color='f2f2f2', end_color='f2f2f2')
            sheet['B' + str(column)] = history_id
            sheet['B' + str(column)].fill = PatternFill(fill_type='solid', start_color='f2f2f2', end_color='f2f2f2')
            sheet['C' + str(column)] = data_count
            sheet['C' + str(column)].fill = PatternFill(fill_type='solid', start_color='f2f2f2', end_color='f2f2f2')
            sheet['D' + str(column)] = data_totalItems
            sheet['D' + str(column)].fill = PatternFill(fill_type='solid', start_color='f2f2f2', end_color='f2f2f2')
            sheet['E' + str(column)] = data_totalDeposit
            sheet['E' + str(column)].fill = PatternFill(fill_type='solid', start_color='f2f2f2', end_color='f2f2f2')
            sheet['F' + str(column)] = history_crash
            sheet['F' + str(column)].fill = PatternFill(fill_type='solid', start_color='f2f2f2', end_color='f2f2f2')
        else:
            sheet['A' + str(column)] = date
            sheet['A' + str(column)].fill = PatternFill(fill_type='solid', start_color='d4d4d4', end_color='d4d4d4')
            sheet['B' + str(column)] = history_id
            sheet['B' + str(column)].fill = PatternFill(fill_type='solid', start_color='d4d4d4', end_color='d4d4d4')
            sheet['C' + str(column)] = data_count
            sheet['C' + str(column)].fill = PatternFill(fill_type='solid', start_color='d4d4d4', end_color='d4d4d4')
            sheet['D' + str(column)] = data_totalItems
            sheet['D' + str(column)].fill = PatternFill(fill_type='solid', start_color='d4d4d4', end_color='d4d4d4')
            sheet['E' + str(column)] = data_totalDeposit
            sheet['E' + str(column)].fill = PatternFill(fill_type='solid', start_color='d4d4d4', end_color='d4d4d4')
            sheet['F' + str(column)] = history_crash
            sheet['F' + str(column)].fill = PatternFill(fill_type='solid', start_color='d4d4d4', end_color='d4d4d4')
        if list_value == 0:
            list_value = 1
        elif list_value == 1:
            list_value = 2
        elif list_value == 2:
            list_value = 1

        excelFileName = 'csgorun_statistic_' + str(list_value) + '.xlsx'
        excelFileName_remove = 'csgorun_statistic_' + str(list_value) + '.xlsx'
        wb.save(excelFileName)

        column += 1
        color_value = not color_value
        last_game_id = history_id


    sleep(8)
